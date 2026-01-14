import requests
import json
import time
from datetime import datetime
import os
import pandas as pd
import yaml

def load_config(config_file: str = "config.yaml") -> dict:
    """加载配置文件"""
    try:
        with open(config_file, 'r', encoding='utf-8') as f:
            return yaml.safe_load(f)
    except Exception as e:
        print(f"❌ 无法读取配置文件 {config_file}: {e}")
        exit(1)

def load_questions(file_path: str) -> tuple:
    """从Excel文件加载测试问题和答案"""
    try:
        df = pd.read_excel(file_path)
        questions = df['问题'].tolist()
        answers = df['答案'].tolist() if '答案' in df.columns else [None] * len(questions)
        return questions, answers
    except Exception as e:
        print(f"❌ 无法读取问题文件 {file_path}: {e}")
        return [], []

# ======================
# 核心函数
# ======================
def ask_model(question: str, config: dict) -> dict:
    """调用本地模型 API"""
    payload = {
        "model": config["model_name"],
        "messages": [{"role": "user", "content": question}],
        "temperature": config["temperature"],
        "max_tokens": config["max_tokens"]
    }
    
    start_time = time.time()
    try:
        response = requests.post(
            config["api_url"],
            json=payload,
            timeout=config["timeout"]
        )
        latency = time.time() - start_time
        
        if response.status_code != 200:
            return {
                "success": False,
                "answer": "",
                "error": f"HTTP {response.status_code}: {response.text}",
                "latency": latency
            }
        
        data = response.json()
        answer = data["choices"][0]["message"]["content"].strip()
        # 移除think过程
        if "<think>" in answer:
            answer = answer.split("</think>")[-1].strip()
        return {
            "success": True,
            "answer": answer,
            "error": "",
            "latency": round(latency, 2)
        }
        
    except Exception as e:
        latency = time.time() - start_time
        return {
            "success": False,
            "answer": "",
            "error": str(e),
            "latency": round(latency, 2)
        }

def calculate_similarity(model_answers: list, reference_answers: list) -> list:
    """计算模型回答与参考答案的相似度（使用简单字符串匹配）"""
    try:
        from difflib import SequenceMatcher
        similarities = []
        
        for i, (model_ans, ref_ans) in enumerate(zip(model_answers, reference_answers)):
            if not ref_ans or pd.isna(ref_ans) or str(ref_ans).strip() == "":
                similarities.append(0.0)
                continue
                
            if not model_ans or model_ans == "":
                similarities.append(0.0)
                continue
                
            # 使用SequenceMatcher计算字符串相似度
            similarity = SequenceMatcher(None, str(model_ans), str(ref_ans)).ratio()
            similarities.append(round(float(similarity), 4))
            
        return similarities
    except Exception as e:
        print(f"❌ 相似度计算失败: {e}")
        return [0.0] * len(model_answers)
def update_questions_file(file_path: str, results: list, similarities: list):
    """将答案和相似度回写到问题Excel文件"""
    try:
        from openpyxl import load_workbook
        from openpyxl.styles import Alignment
        
        df = pd.read_excel(file_path)
        df['回答'] = [r["answer"] for r in results]
        df['相似度'] = similarities
        df.to_excel(file_path, index=False)
        
        # 设置回答列的对齐方式为填充
        wb = load_workbook(file_path)
        ws = wb.active
        answer_col = df.columns.get_loc('回答') + 1
        
        for row in range(2, len(df) + 2):
            ws.cell(row=row, column=answer_col).alignment = Alignment(horizontal='fill')
        
        wb.save(file_path)
        print(f"✅ 答案和相似度已回写到 {file_path}")
    except Exception as e:
        print(f"❌ 回写答案失败: {e}")

def generate_excel_report(results: list, questions: list, config: dict):
    total = len(results)
    success_count = sum(1 for r in results if r["success"])
    avg_latency = round(sum(r["latency"] for r in results) / total, 2)
    
    # 判断是否拒答
    refusal_keywords = ["不知道", "无法回答", "无权限", "不能提供", "未授权", "拒绝"]
    
    # 准备Excel数据
    excel_data = []
    for i, res in enumerate(results):
        is_refusal = res["success"] and any(kw in res["answer"] for kw in refusal_keywords)
        status = "成功" if res["success"] else "失败"
        if is_refusal:
            status += "(拒答)"
        
        excel_data.append({
            "序号": i + 1,
            "问题": questions[i],
            "回答": res["answer"],
            "延迟(秒)": res["latency"],
            "状态": status,
            "错误信息": res["error"] if not res["success"] else ""
        })
    
    # 创建DataFrame
    df = pd.DataFrame(excel_data)
    
    # 保存Excel文件
    os.makedirs("reports", exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    excel_path = f"reports/test_report_{timestamp}.xlsx"
    
    with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
        # 写入详细结果
        df.to_excel(writer, sheet_name='测试结果', index=False)
        
        # 写入汇总信息
        summary_data = {
            "项目": ["测试时间", "API地址", "模型名称", "总问题数", "成功响应", "平均延迟(秒)", "安全拒答数"],
            "值": [
                datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                config["api_url"],
                config["model_name"],
                total,
                f"{success_count}/{total} ({success_count/total:.0%})",
                avg_latency,
                sum(1 for r in results if r["success"] and any(kw in r["answer"] for kw in refusal_keywords))
            ]
        }
        summary_df = pd.DataFrame(summary_data)
        summary_df.to_excel(writer, sheet_name='测试汇总', index=False)
    
    print(f"\n✅ 测试完成！Excel报告已保存至: {excel_path}")
    return excel_path

def run_test(config: dict) -> str:
    """执行测试并返回报告路径"""
    try:
        # 加载测试问题和答案
        questions, reference_answers = load_questions(config["questions_file"])
        if not questions:
            print("❌ 未找到有效的测试问题")
            return None
        
        print(f"🚀 开始自动化测试，共 {len(questions)} 个问题...")
        results = []
        
        for i, question in enumerate(questions, 1):
            # 跳过空问题
            if not question or pd.isna(question) or str(question).strip() == "":
                print(f"[{i}/{len(questions)}] 跳过空问题")
                results.append({
                    "question": "",
                    "answer": "",
                    "success": False,
                    "error": "问题为空",
                    "latency": 0
                })
                continue
                
            print(f"[{i}/{len(questions)}] 提问: {question}")
            result = ask_model(question, config)
            results.append({
                "question": question,
                "answer": result["answer"],
                "success": result["success"],
                "error": result["error"],
                "latency": result["latency"]
            })
            time.sleep(config["sleep_interval"])
        
        # 计算相似度
        model_answers = [r["answer"] for r in results]
        similarities = calculate_similarity(model_answers, reference_answers)
        
        # 回写答案和相似度到问题文件
        update_questions_file(config["questions_file"], results, similarities)
        
        # 生成Excel报告
        excel_file = generate_excel_report(results, questions, config)
        
        # 打印汇总信息
        total = len(results)
        success_count = sum(1 for r in results if r["success"])
        print(f"\n📊 测试汇总: 成功 {success_count}/{total}, 平均延迟 {sum(r['latency'] for r in results)/total:.2f}s")
        
        return excel_file
    except Exception as e:
        print(f"❌ 测试执行失败: {e}")
        return None

# ======================
# 主程序
# ======================
if __name__ == "__main__":
    config = load_config()
    run_test(config)