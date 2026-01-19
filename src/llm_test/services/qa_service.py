import requests
import time
from datetime import datetime
import os
import pandas as pd
from difflib import SequenceMatcher
from openpyxl import load_workbook
from openpyxl.styles import Alignment
import logging

logger = logging.getLogger(__name__)

def load_questions(file_path: str = None, text_input: str = None) -> tuple:
    """从Excel文件或文本输入加载测试问题和答案"""
    try:
        if text_input:
            # 从文本输入加载问题
            questions = [q.strip() for q in text_input.strip().split('\n') if q.strip()]
            answers = [None] * len(questions)
            logger.info(f"从文本输入加载问题成功，共 {len(questions)} 个问题")
            return questions, answers
        elif file_path:
            # 从Excel文件加载问题
            df = pd.read_excel(file_path)
            questions = df['问题'].tolist()
            answers = df['答案'].tolist() if '答案' in df.columns else [None] * len(questions)
            logger.info(f"加载问题文件成功: {file_path}, 共 {len(questions)} 个问题")
            return questions, answers
        else:
            logger.error("未提供问题来源")
            return [], []
    except Exception as e:
        logger.error(f"加载问题失败: {e}", exc_info=True)
        return [], []

def ask_model(question: str, config: dict) -> dict:
    """调用本地模型 API"""
    payload = {
        "model": config["model_name"],
        "messages": [{"role": "user", "content": question}],
        "temperature": config["temperature"],
        "max_tokens": config["max_tokens"]
    }
    
    start_time = time.time()
    try:
        response = requests.post(
            config["api_url"],
            json=payload,
            timeout=config["timeout"]
        )
        latency = time.time() - start_time
        
        if response.status_code != 200:
            error_msg = f"HTTP {response.status_code}: {response.text}"
            logger.error(f"API请求失败: {error_msg}")
            return {
                "success": False,
                "answer": "",
                "error": error_msg,
                "latency": latency
            }
        
        data = response.json()
        answer = data["choices"][0]["message"]["content"].strip()
        if "<think>" in answer:
            answer = answer.split("</think>")[-1].strip()
        return {
            "success": True,
            "answer": answer,
            "error": "",
            "latency": round(latency, 2)
        }
        
    except Exception as e:
        latency = time.time() - start_time
        logger.error(f"API调用异常: {e}", exc_info=True)
        return {
            "success": False,
            "answer": "",
            "error": str(e),
            "latency": round(latency, 2)
        }

def calculate_similarity(model_answers: list, reference_answers: list) -> list:
    """计算模型回答与参考答案的相似度"""
    similarities = []
    for model_ans, ref_ans in zip(model_answers, reference_answers):
        if not ref_ans or pd.isna(ref_ans) or not model_ans:
            similarities.append(0.0)
        else:
            similarity = SequenceMatcher(None, str(model_ans), str(ref_ans)).ratio()
            similarities.append(round(similarity, 4))
    return similarities

def update_questions_file(file_path: str, results: list, similarities: list):
    """将答案和相似度回写到问题Excel文件（仅文件模式）"""
    if not file_path:
        logger.info("手动输入模式，跳过回写问题文件")
        return
    try:
        df = pd.read_excel(file_path)
        df['回答'] = [r["answer"] for r in results]
        df['相似度'] = similarities
        df.to_excel(file_path, index=False)
        
        wb = load_workbook(file_path)
        ws = wb.active
        answer_col = df.columns.get_loc('回答') + 1
        for row in range(2, len(df) + 2):
            ws.cell(row=row, column=answer_col).alignment = Alignment(horizontal='fill')
        wb.save(file_path)
        logger.info(f"答案和相似度已回写到 {file_path}")
    except Exception as e:
        logger.error(f"回写答案失败: {e}", exc_info=True)

def generate_excel_report(results: list, questions: list, config: dict):
    try:
        total = len(results)
        success_count = sum(1 for r in results if r["success"])
        avg_latency = round(sum(r["latency"] for r in results) / total, 2)
        refusal_keywords = ["不知道", "无法回答", "无权限", "不能提供", "未授权", "拒绝"]
        
        excel_data = []
        for i, res in enumerate(results):
            status = "成功" if res["success"] else "失败"
            if res["success"] and any(kw in res["answer"] for kw in refusal_keywords):
                status += "(拒答)"
            excel_data.append({
                "序号": i + 1,
                "问题": questions[i],
                "回答": res["answer"],
                "延迟(秒)": res["latency"],
                "状态": status,
                "错误信息": res["error"] if not res["success"] else ""
            })
        
        os.makedirs("reports", exist_ok=True)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        excel_path = f"reports/test_report_{timestamp}.xlsx"
        
        with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
            pd.DataFrame(excel_data).to_excel(writer, sheet_name='测试结果', index=False)
            summary_data = {
                "项目": ["测试时间", "API地址", "模型名称", "总问题数", "成功响应", "平均延迟(秒)", "安全拒答数"],
                "值": [
                    datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                    config["api_url"],
                    config["model_name"],
                    total,
                    f"{success_count}/{total} ({success_count/total:.0%})",
                    avg_latency,
                    sum(1 for r in results if r["success"] and any(kw in r["answer"] for kw in refusal_keywords))
                ]
            }
            pd.DataFrame(summary_data).to_excel(writer, sheet_name='测试汇总', index=False)
        
        logger.info(f"测试报告已生成: {excel_path}")
        return excel_path
    except Exception as e:
        logger.error(f"生成报告失败: {e}", exc_info=True)
        raise

def run_test_stream(config: dict):
    """执行测试并实时流式输出进度"""
    try:
        question_mode = config.get("question_mode", "file")
        
        if question_mode == "input":
            questions, reference_answers = load_questions(text_input=config.get("questions_text", ""))
        else:
            questions, reference_answers = load_questions(file_path=config.get("questions_file"))
        
        if not questions:
            yield {"type": "error", "message": "未找到有效的测试问题"}
            return
        
        logger.info(f"开始测试，共 {len(questions)} 个问题")
        yield {"type": "start", "total": len(questions)}
        results = []
        
        for i, question in enumerate(questions, 1):
            if not question or pd.isna(question) or str(question).strip() == "":
                results.append({"answer": "", "success": False, "error": "问题为空", "latency": 0})
                yield {"type": "skip", "index": i, "total": len(questions)}
                continue
            
            yield {"type": "question", "index": i, "total": len(questions), "question": question}
            result = ask_model(question, config)
            results.append({
                "answer": result["answer"],
                "success": result["success"],
                "error": result["error"],
                "latency": result["latency"]
            })
            yield {"type": "answer", "index": i, "answer": result["answer"], "success": result["success"], "latency": result["latency"], "error": result["error"]}
            time.sleep(config["sleep_interval"])
        
        similarities = calculate_similarity([r["answer"] for r in results], reference_answers)
        
        # 仅在文件模式下回写
        if question_mode == "file":
            update_questions_file(config.get("questions_file"), results, similarities)
        
        excel_file = generate_excel_report(results, questions, config)
        
        success_count = sum(1 for r in results if r["success"])
        logger.info(f"测试完成: {success_count}/{len(results)} 成功")
        yield {"type": "complete", "success_count": success_count, "total": len(results), "report_path": excel_file}
    except Exception as e:
        logger.error(f"测试执行失败: {e}", exc_info=True)
        yield {"type": "error", "message": str(e)}

def run_test(config: dict) -> str:
    """执行测试并返回报告路径"""
    try:
        question_mode = config.get("question_mode", "file")
        
        if question_mode == "input":
            questions, reference_answers = load_questions(text_input=config.get("questions_text", ""))
        else:
            questions, reference_answers = load_questions(file_path=config.get("questions_file"))
        
        if not questions:
            return None
        
        logger.info(f"开始测试，共 {len(questions)} 个问题")
        results = []
        
        for i, question in enumerate(questions, 1):
            if not question or pd.isna(question) or str(question).strip() == "":
                results.append({"answer": "", "success": False, "error": "问题为空", "latency": 0})
                continue
                
            logger.info(f"[{i}/{len(questions)}] {question}")
            result = ask_model(question, config)
            results.append({
                "answer": result["answer"],
                "success": result["success"],
                "error": result["error"],
                "latency": result["latency"]
            })
            time.sleep(config["sleep_interval"])
        
        similarities = calculate_similarity([r["answer"] for r in results], reference_answers)
        
        # 仅在文件模式下回写
        if question_mode == "file":
            update_questions_file(config.get("questions_file"), results, similarities)
        
        excel_file = generate_excel_report(results, questions, config)
        
        success_count = sum(1 for r in results if r["success"])
        logger.info(f"测试完成: {success_count}/{len(results)} 成功")
        return excel_file
    except Exception as e:
        logger.error(f"测试失败: {e}", exc_info=True)
        return None
